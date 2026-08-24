"""Tests for technical audit report generation."""

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from app.core.storage import LocalFileStorage
from app.modules.audit.service import build_site_audit, inspect_audit_sources


CRAWL_CSV = """Ответ сервера,URL страницы,Title страницы,Meta Description
200,https://example.com/,Home,Description
200,https://example.com/catalog,Same,Repeated
200,https://example.com/catalog/item?color=red,Same,Repeated
404,https://example.com/missing,Missing,Missing
"""
SITEMAP_CSV = """url,Ответ сервера
https://example.com/,200
https://example.com/sitemap-only,200
"""


class SiteAuditTests(unittest.TestCase):
    """Verify full and partial audits stay explicit about source availability."""

    def test_builds_full_audit_from_crawl_and_sitemap(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            storage = LocalFileStorage(root=Path(temporary_directory))
            storage.write_text("crawl/example.csv", CRAWL_CSV)
            storage.write_text("sitemap_parsing/example.csv", SITEMAP_CSV)

            result = build_site_audit("Example", storage=storage)
            workbook = load_workbook(storage.root / result.relative_path, read_only=True, data_only=True)

            self.assertEqual(result.crawl_page_count, 4)
            self.assertEqual(result.sitemap_url_count, 2)
            self.assertEqual(result.checks_skipped, 0)
            self.assertIn("Сравнение sitemap", workbook.sheetnames)
            summary = list(workbook["Сводка"].iter_rows(values_only=True))
            self.assertIn(("Страниц не 200", 1, "рассчитано"), summary)
            self.assertIn(("На сайте, но не в sitemap", 1, "рассчитано"), summary)

    def test_marks_sitemap_checks_as_unavailable_when_only_crawl_exists(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            storage = LocalFileStorage(root=Path(temporary_directory))
            storage.write_text("crawl/example.csv", CRAWL_CSV)

            crawl, sitemap = inspect_audit_sources("Example", storage=storage)
            result = build_site_audit("Example", storage=storage)
            workbook = load_workbook(storage.root / result.relative_path, read_only=True, data_only=True)

            self.assertTrue(crawl.exists)
            self.assertFalse(sitemap.exists)
            self.assertIsNone(result.sitemap_url_count)
            self.assertEqual(result.checks_skipped, 3)
            self.assertNotIn("Сравнение sitemap", workbook.sheetnames)
            summary = list(workbook["Сводка"].iter_rows(values_only=True))
            self.assertIn(("URL в sitemap", None, "нет данных sitemap"), summary)

    def test_rejects_audit_without_crawl_export(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            storage = LocalFileStorage(root=Path(temporary_directory))

            with self.assertRaisesRegex(ValueError, "сначала нужен результат парсинга"):
                build_site_audit("Example", storage=storage)
