"""Regression checks for partial crawl report exports."""

import io
import os
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from app.interfaces.bot.services import (
    _build_xlsx_bytes_from_rows,
    _find_relative_link_issues_csv_path,
    _parse_csv_rows,
    _should_rebuild_partial_crawl_xlsx,
)
from app.interfaces.worker.jobs import (
    _TaskCancellationWithResultError,
    _TaskExecutionResult,
    _resolve_failure_result_payload,
)


class PartialCrawlReportTests(unittest.TestCase):
    """Ensure checkpoint XLSX exports retain relative-link diagnostics."""

    def test_checkpoint_xlsx_has_relative_link_issues_sheet(self) -> None:
        workbook_bytes = _build_xlsx_bytes_from_rows(
            [
                ["Ответ сервера", "URL страницы"],
                ["200", "https://example.com/"],
            ],
            [
                ["Источник", "Исходный href", "Получившийся URL"],
                ["https://example.com/", "catalog", "https://example.com/catalog"],
            ],
        )

        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
        self.assertEqual(workbook.sheetnames, ["crawl_pages", "relative_link_issues"])
        self.assertEqual(
            list(workbook["relative_link_issues"].iter_rows(values_only=True)),
            [
                ("Источник", "Исходный href", "Получившийся URL"),
                ("https://example.com/", "catalog", "https://example.com/catalog"),
            ],
        )

    def test_failure_keeps_latest_checkpoint_payload(self) -> None:
        checkpoint_payload = {
            "pages_crawled": 300,
            "export_files": {
                "csv": "crawl/example.csv",
                "relative_link_issues_csv": "crawl/example.relative_link_issues.csv",
            },
        }

        result = _resolve_failure_result_payload(RuntimeError("RQ timeout"), checkpoint_payload)

        self.assertEqual(result, checkpoint_payload)

    def test_cancellation_result_replaces_checkpoint_payload(self) -> None:
        cancellation_payload = {"export_files": {"xlsx": "crawl/example.xlsx"}}
        error = _TaskCancellationWithResultError(
            "Stopped",
            execution_result=_TaskExecutionResult(result_payload=cancellation_payload),
        )

        result = _resolve_failure_result_payload(error, {"pages_crawled": 300})

        self.assertEqual(result, cancellation_payload)

    def test_adjacent_sidecar_is_used_when_payload_is_missing(self) -> None:
        with TemporaryDirectory() as temp_directory:
            storage_root = Path(temp_directory)
            crawl_directory = storage_root / "crawl"
            crawl_directory.mkdir()
            csv_path = crawl_directory / "example.csv"
            csv_path.write_text("Ответ сервера,URL страницы\n200,https://example.com/\n", encoding="utf-8")
            sidecar_path = crawl_directory / "example.relative_link_issues.csv"
            sidecar_path.write_text(
                "Источник,Исходный href,Получившийся URL\n"
                "https://example.com/,catalog,https://example.com/catalog\n",
                encoding="utf-8",
            )

            found_sidecar = _find_relative_link_issues_csv_path(
                storage_root=storage_root,
                csv_relative_path="crawl/example.csv",
                result_payload=None,
            )

            self.assertEqual(found_sidecar, sidecar_path)
            workbook_bytes = _build_xlsx_bytes_from_rows(
                _parse_csv_rows(csv_path.read_text(encoding="utf-8")),
                _parse_csv_rows(sidecar_path.read_text(encoding="utf-8")),
            )
            workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
            self.assertEqual(
                list(workbook["relative_link_issues"].iter_rows(values_only=True))[1],
                ("https://example.com/", "catalog", "https://example.com/catalog"),
            )

    def test_unchanged_checkpoint_does_not_rebuild_xlsx(self) -> None:
        with TemporaryDirectory() as temp_directory:
            directory = Path(temp_directory)
            csv_path = directory / "example.csv"
            sidecar_path = directory / "example.relative_link_issues.csv"
            xlsx_path = directory / "example.xlsx"
            for path in (csv_path, sidecar_path, xlsx_path):
                path.write_text("content", encoding="utf-8")

            os.utime(csv_path, (100, 100))
            os.utime(sidecar_path, (100, 100))
            os.utime(xlsx_path, (200, 200))

            self.assertFalse(
                _should_rebuild_partial_crawl_xlsx(
                    csv_path=csv_path,
                    relative_link_issues_csv_path=sidecar_path,
                    xlsx_path=xlsx_path,
                )
            )

            os.utime(sidecar_path, (300, 300))
            self.assertTrue(
                _should_rebuild_partial_crawl_xlsx(
                    csv_path=csv_path,
                    relative_link_issues_csv_path=sidecar_path,
                    xlsx_path=xlsx_path,
                )
            )
