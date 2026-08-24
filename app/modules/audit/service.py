"""Build technical site-audit reports from stored crawl and sitemap exports."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from app.core.slug import slugify
from app.core.storage import LocalFileStorage


@dataclass(frozen=True, slots=True)
class AuditSourceState:
    """Availability and freshness of one audit input file."""

    name: str
    path: Path
    exists: bool
    row_count: int | None
    updated_at: datetime | None


@dataclass(frozen=True, slots=True)
class SiteAuditResult:
    """Generated report artifact and its compact summary."""

    relative_path: str
    project_name: str
    crawl_page_count: int
    sitemap_url_count: int | None
    checks_calculated: int
    checks_skipped: int


def inspect_audit_sources(project_name: str, *, storage: LocalFileStorage | None = None) -> tuple[AuditSourceState, AuditSourceState]:
    """Return the current crawl and sitemap input state for a project."""

    storage = storage or LocalFileStorage()
    slug = slugify(project_name)
    return (
        _inspect_csv_source("Краулинг", storage.root / "crawl" / f"{slug}.csv"),
        _inspect_csv_source("Sitemap", storage.root / "sitemap_parsing" / f"{slug}.csv"),
    )


def build_site_audit(project_name: str, *, storage: LocalFileStorage | None = None) -> SiteAuditResult:
    """Create an XLSX audit from the latest stored project exports.

    Crawling is the minimum required source. Sitemap-derived checks are skipped
    when its CSV has not been generated yet, rather than reported as zero.
    """

    storage = storage or LocalFileStorage()
    crawl_source, sitemap_source = inspect_audit_sources(project_name, storage=storage)
    if not crawl_source.exists:
        raise ValueError("Для аудита сначала нужен результат парсинга сайта.")

    crawl_rows = _read_csv_rows(crawl_source.path)
    sitemap_rows = _read_csv_rows(sitemap_source.path) if sitemap_source.exists else []
    report_bytes, summary = _build_report_bytes(
        project_name=project_name,
        crawl_source=crawl_source,
        sitemap_source=sitemap_source,
        crawl_rows=crawl_rows,
        sitemap_rows=sitemap_rows,
    )
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    relative_path = f"audits/{slugify(project_name)}/{timestamp}/audit.xlsx"
    storage.write_bytes(relative_path, report_bytes)
    return SiteAuditResult(
        relative_path=relative_path,
        project_name=project_name,
        crawl_page_count=len(crawl_rows),
        sitemap_url_count=len(sitemap_rows) if sitemap_source.exists else None,
        checks_calculated=summary["checks_calculated"],
        checks_skipped=summary["checks_skipped"],
    )


def _inspect_csv_source(name: str, path: Path) -> AuditSourceState:
    if not path.exists():
        return AuditSourceState(name=name, path=path, exists=False, row_count=None, updated_at=None)
    return AuditSourceState(
        name=name,
        path=path,
        exists=True,
        row_count=len(_read_csv_rows(path)),
        updated_at=datetime.fromtimestamp(path.stat().st_mtime, UTC),
    )


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return [dict(row) for row in csv.DictReader(source)]


def _build_report_bytes(
    *,
    project_name: str,
    crawl_source: AuditSourceState,
    sitemap_source: AuditSourceState,
    crawl_rows: list[dict[str, str]],
    sitemap_rows: list[dict[str, str]],
) -> tuple[bytes, dict[str, int]]:
    from openpyxl import Workbook

    pages_200 = [row for row in crawl_rows if row.get("Ответ сервера", "").strip() == "200"]
    not_200 = [row for row in crawl_rows if row.get("Ответ сервера", "").strip() != "200"]
    with_params = [row for row in crawl_rows if "?" in row.get("URL страницы", "")]
    duplicate_titles = _duplicate_rows(pages_200, "Title страницы")
    duplicate_descriptions = _duplicate_rows(pages_200, "Meta Description")

    summary_rows: list[list[object]] = [
        ["Проект", project_name],
        ["Создан", datetime.now(UTC).isoformat(timespec="seconds")],
        [],
        ["Показатель", "Значение", "Статус"],
        ["Страниц в краулинге", len(crawl_rows), "рассчитано"],
        ["Страниц с ответом 200", len(pages_200), "рассчитано"],
        ["Страниц не 200", len(not_200), "рассчитано"],
        ["URL с параметрами", len(with_params), "рассчитано"],
        ["Строк в дублях Title", len(duplicate_titles), "рассчитано"],
        ["Строк в дублях Description", len(duplicate_descriptions), "рассчитано"],
    ]
    checks_calculated = 5
    checks_skipped = 0

    crawl_urls = {
        row.get("URL страницы", "")
        for row in pages_200
        if row.get("URL страницы") and "?" not in row.get("URL страницы", "")
    }
    if sitemap_source.exists:
        sitemap_urls = {row.get("url", "") for row in sitemap_rows if row.get("url")}
        site_not_in_sitemap = sorted(crawl_urls - sitemap_urls)
        sitemap_not_on_site = sorted(sitemap_urls - crawl_urls)
        summary_rows.extend(
            [
                ["URL в sitemap", len(sitemap_urls), "рассчитано"],
                ["На сайте, но не в sitemap", len(site_not_in_sitemap), "рассчитано"],
                ["В sitemap, но не среди страниц 200", len(sitemap_not_on_site), "рассчитано"],
            ]
        )
        checks_calculated += 3
    else:
        site_not_in_sitemap = []
        sitemap_not_on_site = []
        summary_rows.extend(
            [
                ["URL в sitemap", "", "нет данных sitemap"],
                ["На сайте, но не в sitemap", "", "нет данных sitemap"],
                ["В sitemap, но не среди страниц 200", "", "нет данных sitemap"],
            ]
        )
        checks_skipped += 3

    source_rows = [
        ["Источник", "Состояние", "Строк", "Обновлён"],
        *_source_report_row(crawl_source),
        *_source_report_row(sitemap_source),
        ["Яндекс Вебмастер", "пока не подключён к аудиту", "", ""],
        ["Google Search Console", "пока не подключён к аудиту", "", ""],
    ]

    workbook = Workbook(write_only=True)
    _append_sheet(workbook, "Сводка", summary_rows)
    _append_sheet(workbook, "Статус данных", source_rows)
    _append_dict_rows(workbook, "Не 200", not_200)
    _append_dict_rows(workbook, "URL с параметрами", with_params)
    _append_dict_rows(workbook, "Дубли Title", duplicate_titles)
    _append_dict_rows(workbook, "Дубли Description", duplicate_descriptions)
    if sitemap_source.exists:
        _append_sheet(
            workbook,
            "Сравнение sitemap",
            [["На сайте, но не в sitemap", "В sitemap, но не среди страниц 200"], *zip_longest(site_not_in_sitemap, sitemap_not_on_site)],
        )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), {"checks_calculated": checks_calculated, "checks_skipped": checks_skipped}


def _source_report_row(source: AuditSourceState) -> list[list[object]]:
    if not source.exists:
        return [[source.name, "нет файла", "", ""]]
    updated_at = source.updated_at.isoformat(timespec="seconds") if source.updated_at else ""
    return [[source.name, "готов", source.row_count or 0, updated_at]]


def _duplicate_rows(rows: list[dict[str, str]], column: str) -> list[dict[str, str]]:
    values: dict[str, int] = {}
    for row in rows:
        value = row.get(column, "").strip()
        if value:
            values[value] = values.get(value, 0) + 1
    return [row for row in rows if values.get(row.get(column, "").strip(), 0) > 1]


def _append_sheet(workbook, title: str, rows: list[list[object]] | list[tuple[object, ...]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    for row in rows:
        worksheet.append(list(row))


def _append_dict_rows(workbook, title: str, rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        worksheet.append(["Нет данных"])
        return
    columns = list(rows[0])
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])


def zip_longest(left: list[str], right: list[str]):
    """Yield two comparison columns without importing a heavy dataframe dependency."""

    size = max(len(left), len(right))
    for index in range(size):
        yield (
            left[index] if index < len(left) else "",
            right[index] if index < len(right) else "",
        )
