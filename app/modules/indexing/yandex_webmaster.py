"""File-backed Yandex Webmaster recrawl queues and aggregate report."""

from __future__ import annotations

import csv
import os
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook

from app.core.storage import LocalFileStorage

QUEUE_HEADERS = ("queue_id", "URL")
REPORT_HEADERS = (
    "Проект",
    "Канал",
    "Действие",
    "Количество страниц",
    "Дата",
    "Итоговый статус",
    "Текст ошибки",
)
URL_HEADER_CANDIDATES = {"url", "urls", "url страницы"}


@dataclass(frozen=True, slots=True)
class RecrawlQueueEntry:
    """One queue item, including a hidden stable identity for concurrent updates."""

    queue_id: str
    url: str


def get_recrawl_queue_path(project_slug: str) -> Path:
    """Return the CSV queue path for one project, creating parent directories."""

    storage = LocalFileStorage()
    path = storage.root / "indexing" / "yandex_webmaster" / "recrawl" / f"{project_slug}.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def get_recrawl_queue_count(project_slug: str) -> int:
    """Return the number of currently pending URLs in a project queue."""

    return len(_read_queue(get_recrawl_queue_path(project_slug)))


def prepend_recrawl_urls(project_slug: str, urls: list[str]) -> int:
    """Put received URLs before the existing queue without removing duplicates."""

    return _add_urls(project_slug, urls, prepend=True)


def append_recrawl_urls(project_slug: str, urls: list[str]) -> int:
    """Put received URLs after the existing queue without removing duplicates."""

    return _add_urls(project_slug, urls, prepend=False)


def replace_recrawl_urls(project_slug: str, urls: list[str]) -> int:
    """Replace a project's pending recrawl queue with sitemap URLs."""

    queue_path = get_recrawl_queue_path(project_slug)
    entries = [
        RecrawlQueueEntry(queue_id=uuid.uuid4().hex, url=url.strip())
        for url in urls
        if url.strip()
    ]
    with _locked_file(queue_path):
        _write_queue(queue_path, entries)
    return len(entries)


def peek_recrawl_queue(project_slug: str, *, limit: int) -> list[RecrawlQueueEntry]:
    """Read a small current queue slice without claiming or removing it."""

    queue_path = get_recrawl_queue_path(project_slug)
    with _locked_file(queue_path):
        return _read_queue(queue_path)[:limit]


def remove_recrawl_queue_entry(project_slug: str, queue_id: str) -> bool:
    """Remove one accepted entry, preserving duplicates and new priority URLs."""

    queue_path = get_recrawl_queue_path(project_slug)
    with _locked_file(queue_path):
        entries = _read_queue(queue_path)
        updated_entries = [entry for entry in entries if entry.queue_id != queue_id]
        if len(updated_entries) == len(entries):
            return False
        _write_queue(queue_path, updated_entries)
        return True


def append_indexing_report(
    *,
    project_name: str,
    channel: str,
    action: str,
    page_count: int,
    status: str,
    error_text: str | None,
) -> Path:
    """Append one aggregate indexing operation row to the shared XLSX report."""

    storage = LocalFileStorage()
    report_path = storage.root / "reports" / "indexing_report.xlsx"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with _locked_file(report_path):
        if report_path.exists():
            workbook = load_workbook(report_path)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "indexing_report"
            worksheet.append(REPORT_HEADERS)
        worksheet.append(
            [
                project_name,
                channel,
                action,
                page_count,
                datetime.now().astimezone().date().isoformat(),
                status,
                error_text or "",
            ]
        )
        _atomic_save(workbook, report_path)
    return report_path


def _add_urls(project_slug: str, urls: list[str], *, prepend: bool) -> int:
    """Mutate one queue under a cross-process lock."""

    accepted_urls = [url.strip() for url in urls if url.strip()]
    if not accepted_urls:
        return 0
    queue_path = get_recrawl_queue_path(project_slug)
    new_entries = [RecrawlQueueEntry(queue_id=uuid.uuid4().hex, url=url) for url in accepted_urls]
    with _locked_file(queue_path):
        existing_entries = _read_queue(queue_path)
        _write_queue(queue_path, new_entries + existing_entries if prepend else existing_entries + new_entries)
    return len(new_entries)


def _read_queue(queue_path: Path) -> list[RecrawlQueueEntry]:
    """Read a human-managed CSV queue and accept legacy URL header names."""

    if not queue_path.exists():
        return []

    with queue_path.open("r", encoding="utf-8-sig", newline="") as queue_file:
        reader = csv.reader(queue_file)
        header_row = next(reader, None)
        if not header_row:
            return []
        headers = [value.strip().lower() for value in header_row]
        queue_id_index = headers.index("queue_id") if "queue_id" in headers else None
        url_index = next((index for index, header in enumerate(headers) if header in URL_HEADER_CANDIDATES), None)
        if url_index is None:
            raise ValueError(f"В файле {queue_path.name} не найдена колонка URL.")
        entries: list[RecrawlQueueEntry] = []
        for row in reader:
            if url_index >= len(row):
                continue
            url = row[url_index].strip()
            if not url:
                continue
            queue_id = row[queue_id_index].strip() if queue_id_index is not None and queue_id_index < len(row) else ""
            entries.append(RecrawlQueueEntry(queue_id=queue_id or uuid.uuid4().hex, url=url))
    return entries


def _write_queue(queue_path: Path, entries: list[RecrawlQueueEntry]) -> None:
    """Atomically write a queue CSV with stable identifiers for concurrent updates."""

    with NamedTemporaryFile(
        dir=queue_path.parent,
        suffix=".csv",
        mode="w",
        encoding="utf-8",
        newline="",
        delete=False,
    ) as temp_file:
        temp_path = Path(temp_file.name)
        writer = csv.writer(temp_file)
        writer.writerow(QUEUE_HEADERS)
        writer.writerows((entry.queue_id, entry.url) for entry in entries)
    try:
        os.replace(temp_path, queue_path)
    finally:
        temp_path.unlink(missing_ok=True)


def _atomic_save(workbook: Workbook, target_path: Path) -> None:
    """Save workbooks via replace so readers never see a partial XLSX file."""

    with NamedTemporaryFile(dir=target_path.parent, suffix=".xlsx", delete=False) as temp_file:
        temp_path = Path(temp_file.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, target_path)
    finally:
        workbook.close()
        temp_path.unlink(missing_ok=True)


class _locked_file:
    """Cross-process advisory lock held only during queue/report file mutations."""

    def __init__(self, target_path: Path) -> None:
        self._lock_path = target_path.with_suffix(f"{target_path.suffix}.lock")
        self._handle = None

    def __enter__(self) -> None:
        import fcntl

        self._handle = self._lock_path.open("a+")
        fcntl.flock(self._handle.fileno(), fcntl.LOCK_EX)

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        import fcntl

        if self._handle is not None:
            fcntl.flock(self._handle.fileno(), fcntl.LOCK_UN)
            self._handle.close()
